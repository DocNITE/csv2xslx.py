import pandas as pd
import openpyxl
import csv
import sys
import os

from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

"""
Данный скрипт помогает импортировать CSV файлы в единый XSLS, 
с использованием ТОЛЬКО суммарной информации.

Page,Device(No use!),Group,ItemID(Not use!),Item,Value

TODO: Сделать игнор для Ввод, DMI, некоторых полей.
Затем добавить автотитульник от "Имя компьютера"
Также название для sheet должно идти от "Имя компьютера"
"""

SUM_INFO_DEF = "Суммарная информация"
PC_NAME_FIELD_DEF = "Имя компьютера"
FONT_DEF = "Times New Roman"

ROWS_MAX_DEF = 256
MAIN_TITLE_SIZE_DEF = 20
TITLE_SIZE_DEF = 14
CONTENT_SIZE_DEF = 12
A_ROW_WIDTH_DEF = 11
B_ROW_WIDTH_DEF = 30
C_ROW_WIDTH_DEF = 50
MAX_PRINTERS = 4
MAX_GPU = 1

IGNORE_TITLES = ["DMI", "Ввод"]
IGNORE_ITEMS = [
"DirectX", "Edge", "Internet Explorer", 
"Дата / Время", "SMART-статус жёстких дисков", 
"Контроллер хранения данных", "Вход в домен",
"Коммуникационный порт"]

PINTERS_DEF = "Принтер"
GPU_DEF = "Видеоадаптер"

FILE_NAME_DEF = sys.argv[1]
FILES_DIR_DEF = "./"

"""
Проверяет является ли массив частью "Суммарной информации"
Возвращает сконвертированный его вариант
"""
def isSumInfo(arr):
    if arr[0] != SUM_INFO_DEF:
        return False
    # Group, Item, Value
    new_arr = [arr[2], arr[4], arr[5]]
    return new_arr

"""
Возвращает всю информацию о файле для импорта в .xslx
 # ri - счестчик строк
 # arr_info - сконвертированный массив с нужным данными
 # result - возвращаемая инфа о файле
"""
def getFileInfo(file_name):
    with open (file_name, "r") as file:
        reader = csv.reader(file)
        result = [] # читаем это в writeExcel
        ri = 0
        for row in reader:
            ri=ri+1
            if ri > ROWS_MAX_DEF:
                break
            arr_info = isSumInfo(row)
            if arr_info == False:
                continue
            result.append(arr_info)
        return [file_name, result]
        
"""
yeas
"""
def checkIgnoreTitle(title):
    for data in IGNORE_TITLES:
        if data == title:
            return True
    return False

def checkIgnoreItem(item):
    for data in IGNORE_ITEMS:
        if data == item:
            return True
    return False

"""
Конвертирует полученный массив в excel data для таблицы.
"""
def toFrame(data):
    result = []
    result.append(["", "", ""])
    pc_name = ""
    old_group = ""
    curr_printers = 0
    curr_gpu = 0
    for arr in data:
        print(arr)
        print("\n")
        if old_group != arr[0] and checkIgnoreTitle(arr[0]) != True:
            result.append([arr[0]+":", "", ""])
            old_group = arr[0]
            # continue
        if arr[1] == PC_NAME_FIELD_DEF:
            pc_name = arr[2]
        if arr[1] == PINTERS_DEF:
            curr_printers = curr_printers + 1
        if arr[1] == GPU_DEF:
            curr_gpu = curr_gpu + 1
        # shit statement
        if checkIgnoreTitle(arr[0]) != True:
            if checkIgnoreItem(arr[1]) != True:
                if arr[1] != PINTERS_DEF and arr[2] != GPU_DEF:
                    result.append(["", arr[1], arr[2]])
                elif arr[1] == PINTERS_DEF and curr_printers <= MAX_PRINTERS:
                    result.append(["", arr[1], arr[2]])
                elif arr[1] == GPU_DEF and curr_gpu <= MAX_GPU:
                    result.append(["", arr[1], arr[2]])
    return [pd.DataFrame(result), pc_name]

"""
Записывает данные в excel. Also - добавляет стили и все такое
"""
def writeExcel(data, name):
    with pd.ExcelWriter(name, engine="openpyxl") as writer:
        for data_f in data:
            # writing
            full_fc = toFrame(data_f[1])
            pc_name = full_fc[1]
            # sheet_n = data_f[0] # todo: shoulde be use PC name from report
            sheet_n = pc_name
            data_fc = full_fc[0]
            data_fc.to_excel(writer, sheet_name=sheet_n, index=False)
            #todo: also - add title support from pc name 
            # styling
            workbook = writer.book
            sheet = workbook[sheet_n]
            sheet.column_dimensions['A'].width = A_ROW_WIDTH_DEF
            sheet.column_dimensions['B'].width = B_ROW_WIDTH_DEF
            sheet.column_dimensions['C'].width = C_ROW_WIDTH_DEF
            border_style = Side(border_style="thick")
            border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
            border_title = Border(bottom=border_style)
            border_nothing = Border(top=border_style, bottom=border_style)
            alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            alignment_a = Alignment(horizontal='left')
            alignment_title = Alignment(horizontal='center')
            # formating B column 
            for cell in sheet['B']:
                cell.font = Font(name=FONT_DEF, size=CONTENT_SIZE_DEF)
                cell.border = border
                cell.alignment = alignment
            # formating C column 
            for cell in sheet['C']:
                cell.font = Font(name=FONT_DEF, size=CONTENT_SIZE_DEF)
                cell.border = border
                cell.alignment = alignment
            # formating A column 
            ai = 1      # A cells iterate
            s_ai = 0    # start empty A cells iterate
            for cell in sheet['A']:
                if cell.value != "":
                    if s_ai != 0:
                        # merge for vertical cells 
                        sheet.merge_cells("A"+str(s_ai)+":A"+str(ai-1))
                        # restore start A interate
                        s_ai = 0
                    # merge for current horizontal title line
                    sheet.merge_cells("A"+str(ai)+":C"+str(ai))
                else:
                    if s_ai == 0:
                        s_ai = ai
                cell.font = Font(name=FONT_DEF, size=TITLE_SIZE_DEF)
                cell.border = border
                cell.alignment = alignment_a
                ai = ai + 1
            if s_ai != 0:
                sheet.merge_cells("A"+str(s_ai)+":A"+str(ai-1))
                s_ai = 0
            sheet['A1'].value = pc_name
            sheet.merge_cells('A1:C1')
            sheet['A1'].font = Font(name=FONT_DEF, size=MAIN_TITLE_SIZE_DEF)
            # borders
            sheet['A1'].border = border_title
            sheet['B1'].border = border_title
            sheet['C1'].border = border_title
            sheet['A2'].border = border_nothing
            sheet['A2'].font = Font(name=FONT_DEF, size=MAIN_TITLE_SIZE_DEF)
            sheet['B2'].border = border_nothing
            sheet['C2'].border = border_nothing
            sheet.merge_cells('A2:C2')
            # alignments 
            sheet['A1'].alignment = alignment_title
            workbook.save(name)
           
"""
Ищет все нужные файлы для бандла
"""
def getAllFiles(folder_path):
    extension = ".csv"
    file_list = []
    for file in os.listdir(folder_path):
        if file.endswith(extension):
            file_list.append(file)
    return file_list
           
"""
Это база
"""
def main():
    file_arrs = []
    for file in getAllFiles(FILES_DIR_DEF):
        file_arrs.append(getFileInfo(file))
        
    if len(file_arrs) == 0:
        print("Can't find .csv files!")
        return True
        
    writeExcel(file_arrs, FILE_NAME_DEF)
    print("\n Done! "+ FILE_NAME_DEF +" was created!")
main()